"""
exporter.py – Write extracted invoice data to a styled Excel workbook.

Two sheets per output file:
  Sheet 1 "Invoice Summary" – one row per invoice, all scalar fields
  Sheet 2 "Line Items"      – all line items across invoices, with source file column

Uses openpyxl for styling (header colors, borders, auto column widths).
"""

from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.ocr_engine.field_extractor import ExtractedInvoice, LineItem
from src.utils.config import OUTPUT_DIR
from src.utils.logger import get_logger

log = get_logger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Style constants
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
ALT_ROW_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
BORDER_SIDE   = Side(style="thin", color="B8B8B8")
CELL_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")

# Column order for Summary sheet
SUMMARY_COLUMNS = [
    ("source_file",    "Source File"),
    ("ocr_confidence", "OCR Confidence"),
    ("invoice_number", "Invoice #"),
    ("invoice_date",   "Invoice Date"),
    ("due_date",       "Due Date"),
    ("purchase_order", "PO Number"),
    ("order_id",       "Order ID"),
    ("ship_mode",      "Ship Mode"),
    ("vendor_name",    "Vendor Name"),
    ("vendor_address", "Vendor Address"),
    ("vendor_gstin",   "Vendor GSTIN"),
    ("vendor_pan",     "Vendor PAN"),
    ("buyer_name",     "Buyer Name"),
    ("buyer_address",  "Buyer Address"),
    ("currency",       "Currency"),
    ("subtotal",       "Subtotal"),
    ("tax_amount",     "Tax Amount"),
    ("total_amount",   "Grand Total"),
]

# Column order for Line Items sheet
LINE_ITEM_COLUMNS = [
    ("source_file",   "Source File"),
    ("invoice_number","Invoice #"),
    ("sr_no",         "Sr. No."),
    ("description",   "Description"),
    ("hsn_sac_code",  "HSN/SAC"),
    ("quantity",      "Qty"),
    ("unit_price",    "Unit Price"),
    ("discount",      "Disc. %"),
    ("tax_rate",      "Tax Rate"),
    ("tax_amount",    "Tax Amount"),
    ("cgst_rate",     "CGST %"),
    ("cgst_amount",   "CGST Amt"),
    ("sgst_rate",     "SGST %"),
    ("sgst_amount",   "SGST Amt"),
    ("line_total",    "Line Total"),
]


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _write_header(ws, columns: list[tuple[str, str]]) -> None:
    """Write a styled header row to a worksheet."""
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_data_row(ws, row_idx: int, data: dict, columns: list[tuple[str, str]]) -> None:
    """Write a single data row with alternating background."""
    fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
    for col_idx, (key, _) in enumerate(columns, start=1):
        value = data.get(key, "")
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill      = fill
        cell.border    = CELL_BORDER
        cell.alignment = WRAP_ALIGN
        
        # Flag low OCR confidence with red fill
        if key == "ocr_confidence" and isinstance(value, str) and value.endswith("%"):
            try:
                conf_val = float(value.strip("%"))
                if conf_val < 80.0:
                    # Light red fill for low confidence
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006")
            except ValueError:
                pass


def _auto_column_width(ws) -> None:
    """Set column widths based on max content length (capped at 40)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                max_len  = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _build_dynamic_columns(base_columns: List[tuple], data_list: List[dict]) -> List[tuple]:
    """
    1. Finds any extra keys in data_list not present in base_columns.
    2. Removes any columns that are entirely empty across all rows.
    """
    if not data_list:
        return base_columns

    # Collect all keys present across all dicts
    all_keys = set()
    for d in data_list:
        all_keys.update(d.keys())

    # Map existing base keys
    base_keys = {k for k, v in base_columns}

    # Append new keys
    extra_cols = []
    for k in all_keys:
        if k not in base_keys:
            # Format title e.g. "salesperson_name" -> "Salesperson Name"
            extra_cols.append((k, k.replace("_", " ").title()))
    
    full_columns = base_columns + extra_cols

    # Prune empty columns
    active_columns = []
    for col_key, col_label in full_columns:
        # Check if at least one row has a truthy value for this column
        is_empty = all(not str(d.get(col_key) or "").strip() for d in data_list)
        if not is_empty or col_key in ("invoice_number", "source_file"):
            active_columns.append((col_key, col_label))

    return active_columns


def _build_line_item_columns(data_list: List[dict]) -> List[tuple[str, str]]:
    """
    Dynamically discover all columns in data_list and order them in the exact
    raw insertion order of key occurrences.
    """
    if not data_list:
        return [("source_file", "Source File"), ("invoice_number", "Invoice #")]
        
    cols = [("source_file", "Source File"), ("invoice_number", "Invoice #")]
    
    # Collect unique keys while preserving order
    seen_keys = {"source_file", "invoice_number"}
    ordered_keys = []
    for d in data_list:
        for k in d.keys():
            if k not in seen_keys:
                seen_keys.add(k)
                ordered_keys.append(k)
                
    for k in ordered_keys:
        cols.append((k, str(k).strip()))
        
    # Prune empty columns
    active_cols = []
    for col_key, col_label in cols:
        is_empty = all(not str(d.get(col_key) or "").strip() for d in data_list)
        if not is_empty:
            active_cols.append((col_key, col_label))
            
    return active_cols


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    invoices: List[ExtractedInvoice],
    output_path: str | Path | None = None,
) -> Path:
    """
    Export a list of extracted invoices to a styled Excel workbook.

    Args:
        invoices:    List of ExtractedInvoice objects.
        output_path: Optional explicit output path. Defaults to output/excel/invoices.xlsx.

    Returns:
        Path to the created Excel file.
    """
    if not invoices:
        raise ValueError("No invoices to export.")

    if output_path is None:
        output_path = OUTPUT_DIR / "invoices.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log.info("Exporting {} invoice(s) to {}", len(invoices), output_path)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Invoice Summary"
    ws_summary.freeze_panes = "A2"   # freeze header row

    # 1. Collect all summary data
    summary_data_list = [inv.to_flat_dict() for inv in invoices]
    
    # 2. Discover extra keys and drop empty columns
    dynamic_summary_cols = _build_dynamic_columns(SUMMARY_COLUMNS, summary_data_list)
    
    _write_header(ws_summary, dynamic_summary_cols)

    for row_idx, data in enumerate(summary_data_list, start=2):
        _write_data_row(ws_summary, row_idx, data, dynamic_summary_cols)

    _auto_column_width(ws_summary)

    # ── Sheet 2: Line Items ───────────────────────────────────────────────────
    ws_items = wb.create_sheet("Line Items")
    ws_items.freeze_panes = "A2"

    # 1. Collect all line item data dynamically preserving raw columns
    line_item_data_list = []
    for invoice in invoices:
        for item in invoice.line_items:
            data = {
                "source_file"   : invoice.source_file.name,
                "invoice_number": invoice.invoice_number or "",
            }
            if item.extra_fields:
                data.update(item.extra_fields)
            else:
                data.update({
                    "Sr. No."    : item.sr_no,
                    "Description": item.description,
                    "HSN/SAC"    : item.hsn_sac_code,
                    "Qty"        : item.quantity,
                    "Unit Price" : item.unit_price,
                    "Disc. %"    : item.discount,
                    "Tax Rate"   : item.tax_rate,
                    "Tax Amount" : item.tax_amount,
                    "CGST %"     : item.cgst_rate,
                    "CGST Amt"   : item.cgst_amount,
                    "SGST %"     : item.sgst_rate,
                    "SGST Amt"   : item.sgst_amount,
                    "Line Total" : item.line_total,
                })
            line_item_data_list.append(data)

    # 2. Discover raw columns order and drop empty columns
    dynamic_item_cols = _build_line_item_columns(line_item_data_list)

    _write_header(ws_items, dynamic_item_cols)

    for row_idx, data in enumerate(line_item_data_list, start=2):
        _write_data_row(ws_items, row_idx, data, dynamic_item_cols)

    if not line_item_data_list:
        ws_items.cell(row=2, column=1, value="No line items extracted.")

    _auto_column_width(ws_items)

    try:
        wb.save(output_path)
        log.info("Excel workbook saved → {}", output_path)
        return output_path
    except PermissionError:
        # If the file is open in Excel, Windows locks it. 
        # Fallback to appending a timestamp to the filename.
        import time
        timestamp = int(time.time())
        fallback_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        wb.save(fallback_path)
        log.warning("Original file was locked (likely open in Excel). Saved as fallback → {}", fallback_path)
        return fallback_path
